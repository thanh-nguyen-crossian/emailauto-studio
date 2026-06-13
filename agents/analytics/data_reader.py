import json
import re
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from functools import lru_cache
from openpyxl import load_workbook

SOURCE_DIR = Path(__file__).parent.parent.parent / "Source"
MASTER_PLAN_PATH = SOURCE_DIR / "RMKT Master Plan.xlsx"

BRAND_MAP = {
    "bragoddess": "Bra Goddess",
    "gentslux": "GentsLux",
    "luxfitting": "LuxFitting",
    "santafare": "SantaFare",
}

CORE_WORKBOOK_SHEETS = {
    "2026 Forecast",
    "2026 Adjusted Forecast",
    "KPO 2026",
    "Rev-KPO 2026",
    "Timeline",
    "Data Planning",
    "Email Temp Analysis",
    "Schedule",
    "Bra Goddess Metrics",
    "GentsLux Metrics",
    "LuxFitting Metrics",
    "SantaFare Metrics",
    "Segmentation",
    "Email Content",
}


def _month_key(value) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.strftime("%Y-%m")


def _within_month_range(month: str | None, start_month: str | None, end_month: str | None) -> bool:
    if not month:
        return True
    if start_month and month < start_month:
        return False
    if end_month and month > end_month:
        return False
    return True


@lru_cache(maxsize=1)
def list_master_plan_sheets() -> list[str]:
    """Return every sheet in the RMKT workbook without loading cell data."""
    wb = load_workbook(MASTER_PLAN_PATH, read_only=True, data_only=True)
    return list(wb.sheetnames)


def workbook_sheet_catalog() -> list[dict]:
    return [
        {
            "name": name,
            "recommended": name in CORE_WORKBOOK_SHEETS,
        }
        for name in list_master_plan_sheets()
    ]


def extract_js_data_object(html_content: str, var_name: str = "DATA") -> dict:
    """Extract an embedded JS object using the JSON decoder's own boundary detection."""
    pattern = rf"const\s+{var_name}\s*=\s*\{{"
    match = re.search(pattern, html_content)
    if not match:
        raise ValueError(f"No 'const {var_name} =' found in HTML")
    start = match.end() - 1  # rewind to opening brace
    obj, _ = json.JSONDecoder().raw_decode(html_content, start)
    return obj


def read_rmkt_email_dashboard() -> dict:
    path = SOURCE_DIR / "rmkt_email_dashboard.html"
    return extract_js_data_object(path.read_text(encoding="utf-8"))


def read_rmkt_dashboard() -> dict:
    """Extract customer/RFM analytics from rmkt_dashboard.html.

    This file uses named consts (seg, ltv, email, retp, rets) instead of a single DATA object.
    """
    path = SOURCE_DIR / "rmkt_dashboard.html"
    html = path.read_text(encoding="utf-8")
    return {
        "seg": extract_js_data_object(html, "seg"),
        "ltv": extract_js_data_object(html, "ltv"),
        "email": extract_js_data_object(html, "email"),
        "retp": extract_js_data_object(html, "retp"),
        "rets": extract_js_data_object(html, "rets"),
    }


def read_welcome_flow_dashboard() -> dict:
    path = SOURCE_DIR / "welcome_flow_dashboard.html"
    return extract_js_data_object(path.read_text(encoding="utf-8"))


def read_winback_flow_dashboard() -> dict:
    path = SOURCE_DIR / "winback_flow_dashboard.html"
    return extract_js_data_object(path.read_text(encoding="utf-8"))


def _read_optional(reader, default):
    try:
        return reader()
    except FileNotFoundError:
        return default


@lru_cache(maxsize=1)
def read_master_plan_monthly() -> list[dict]:
    """Read historical monthly aggregate data from RMKT Master Plan 2026 Forecast sheet.

    Header is at row index 1 (row 0 is the sheet title '2026 Forecast').
    Rows with Month == 'Total', '2026 Forecast', or NaN are skipped.
    """
    df = pd.read_excel(MASTER_PLAN_PATH, sheet_name="2026 Forecast", engine="openpyxl", skiprows=1)
    df = df.rename(columns=lambda c: str(c).strip())
    # Drop non-data rows
    df = df[
        df["Month"].notna()
        & (~df["Month"].astype(str).isin(["Total", "2026 Forecast", "nan"]))
    ]
    df["Month"] = pd.to_datetime(df["Month"], errors="coerce").dt.strftime("%Y-%m")
    df = df[df["Month"].notna()]
    keep_cols = [
        "Month", "Delivered", "CBH", "PO", "AOV", "Open", "Access", "Optout", "Spam",
        "Open/Delivered", "Access/Delivered", "Access/Open", "CBH/Delivered",
    ]
    available = [c for c in keep_cols if c in df.columns]
    return df[available].to_dict("records")


@lru_cache(maxsize=1)
def read_master_plan_timeline_bounds() -> dict:
    rows = read_master_plan_monthly()
    months = [
        row.get("Month")
        for row in rows
        if row.get("Month") and not pd.isna(row.get("Access/Delivered"))
    ]
    months = sorted(set(months))
    return {
        "min_month": months[0] if months else "",
        "max_month": months[-1] if months else "",
        "months": months,
    }


@lru_cache(maxsize=1)
def read_analysis_timeline_bounds() -> dict:
    rmkt_email = _read_optional(read_rmkt_email_dashboard, {})
    months = []
    for raw_key in BRAND_MAP.values():
        brand_data = rmkt_email.get(raw_key, {})
        for row in brand_data.get("monthly", []):
            month = _month_key(row.get("month"))
            if month and row.get("Delivered"):
                months.append(month)
    if not months:
        first_of_month = datetime.now().replace(day=1)
        last_completed_month = (first_of_month - timedelta(days=1)).strftime("%Y-%m")
        months = [
            row.get("Month")
            for row in read_master_plan_monthly()
            if row.get("Month") and row.get("Month") <= last_completed_month
        ]
    months = sorted(set(months))
    return {
        "min_month": months[0] if months else "",
        "max_month": months[-1] if months else "",
        "months": months,
    }


def read_page_performance() -> list[dict]:
    """Read all 'Page performance *.csv' files from SOURCE_DIR.

    Each CSV starts with a BOM + 'sep=,' line that must be skipped (skiprows=1).
    Filters to RMKT publisher_type rows only.
    """
    records = []
    for csv_file in sorted(SOURCE_DIR.glob("Page performance *.csv")):
        df = pd.read_csv(csv_file, skiprows=1, encoding="utf-8-sig")
        df = df[df["publisher_type"].fillna("").str.strip() == "RMKT"]
        df = df[df["stats_purchase"].notna() & (df["stats_purchase"] > 0)]
        df["source_month"] = csv_file.stem.replace("Page performance ", "")
        keep = [
            "cta_url_domain", "pbase_line", "page_url",
            "stats_access", "stats_purchase", "stats_revenue",
            "stats_addtocart", "stats_initcheckout", "source_month",
        ]
        available = [c for c in keep if c in df.columns]
        records.extend(df[available].to_dict("records"))
    return records


def _recompute_group_stats(rows: list[dict], field: str) -> list[dict]:
    groups: dict[str, list[dict]] = {}
    for row in rows:
        key = str(row.get(field) or "").strip()
        if not key:
            continue
        groups.setdefault(key, []).append(row)

    output = []
    for key, items in groups.items():
        access_rates = [float(r.get("access_rate")) for r in items if r.get("access_rate") is not None]
        cbh_rates = [float(r.get("cbh_per_1k")) for r in items if r.get("cbh_per_1k") is not None]
        output.append({
            field.replace("content_type", "type"): key,
            "segment" if field == "segment" else "type": key,
            "n_sends": len(items),
            "avg_access_rate": sum(access_rates) / len(access_rates) if access_rates else 0,
            "avg_cbh_1k": sum(cbh_rates) / len(cbh_rates) if cbh_rates else 0,
        })
    return sorted(output, key=lambda item: (item.get("avg_access_rate") or 0, item.get("n_sends") or 0), reverse=True)


def apply_timeline_filter(metrics: dict, start_month: str | None = None, end_month: str | None = None) -> dict:
    if not start_month and not end_month:
        return metrics

    for data in metrics.get("brands", {}).values():
        data["monthly"] = [
            row for row in data.get("monthly", [])
            if _within_month_range(_month_key(row.get("month")), start_month, end_month)
        ]
        data["sends"] = [
            row for row in data.get("sends", [])
            if _within_month_range(_month_key(row.get("send_date") or row.get("month")), start_month, end_month)
        ]
        data["content_types"] = _recompute_group_stats(data.get("sends", []), "content_type")
        data["segments"] = _recompute_group_stats(data.get("sends", []), "segment")

    metrics["master_plan_monthly"] = [
        row for row in metrics.get("master_plan_monthly", [])
        if _within_month_range(_month_key(row.get("Month")), start_month, end_month)
    ]
    metrics["page_performance"] = [
        row for row in metrics.get("page_performance", [])
        if _within_month_range(_month_key(row.get("source_month")), start_month, end_month)
    ]
    return metrics


def build_metrics(
    *,
    selected_sources: list[str] | None = None,
    selected_sheets: list[str] | None = None,
    start_month: str | None = None,
    end_month: str | None = None,
) -> dict:
    """Read all data sources and return a normalized metrics dict.

    This is the single entry point for the data layer.
    All downstream components (metrics_engine, analyst, report_generator)
    consume the dict produced here.
    """
    rmkt_email = _read_optional(read_rmkt_email_dashboard, {})
    rmkt = _read_optional(read_rmkt_dashboard, {})
    welcome = _read_optional(read_welcome_flow_dashboard, {})
    winback = _read_optional(read_winback_flow_dashboard, {})

    brands = {}
    for slug, raw_key in BRAND_MAP.items():
        brand_data = rmkt_email.get(raw_key, {})
        brands[slug] = {
            "monthly": brand_data.get("monthly", []),
            "sends": brand_data.get("sends", []),
            "kpi": brand_data.get("kpi", {}),
            "content_types": brand_data.get("content_types", []),
            "segments": brand_data.get("segments", []),
            "target": brand_data.get("target", {}),
            "factor_table": brand_data.get("factor_table", []),
            "flow_health": {
                "welcome": welcome.get(raw_key, {}),
                "winback": winback.get(raw_key, {}),
            },
        }

    selected_sources = selected_sources or ["master_plan"]
    selected_sheets = selected_sheets or list_master_plan_sheets()
    page_performance = read_page_performance() if "page_performance" in selected_sources else []

    metrics = {
        "generated_at": datetime.now().isoformat(),
        "brands": brands,
        "customers": rmkt,
        "master_plan_monthly": read_master_plan_monthly(),
        "page_performance": page_performance,
        "analysis_scope": {
            "sources": selected_sources,
            "workbook": MASTER_PLAN_PATH.name,
            "sheets": selected_sheets,
            "start_month": start_month or "",
            "end_month": end_month or "",
        },
        "anomalies": [],
    }
    return apply_timeline_filter(metrics, start_month, end_month)
